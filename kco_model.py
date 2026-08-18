# ============================================================
# KCO — Kuznetsov-Cunningham-Ouchterlony fragmentation model
#       (post-blast fragment-size distribution prediction)
#
# Pure-physics module: no file I/O, no plotting.
#
# Baseline formulation: Ouchterlony (2005), "The Swebrec function:
# linking fragmentation by blasting and crushing", eqs. (11a)-(11e).
# Later developments (Cunningham 2005 timing factor A_T; the shifted
# Kuz-Ram g(n)) are OPTIONAL and DISABLED by default. The A/6 rock-factor
# term in the uniformity index is part of the baseline Cunningham (2005)
# Eq. (48) implementation. Each function documents which published
# formulation it implements.
#
# Workflow:
#   rock/blast inputs -> BI -> A -> X50
#   blast geometry    -> n
#   DFN / in-situ block information -> Xmax
#   X50, Xmax, n      -> b
#   X50, Xmax, b      -> P(x)   (Swebrec cumulative passing)
#
# Units convention (see also BlastDesign):
#   B, S, H, W, Lb, Lc, Ltot   m
#   D (hole diameter)          mm
#   Q                          kg/hole
#   q                          kg/m3
#   rho                        kg/m3
#   UCS (sigma_c)              MPa
#   E (Young's modulus)        GPa
#   s_ANFO                     %
#   X50, Xmax, P(x) sizes      mm in all final outputs
#                              (X50 is computed in cm by the Kuznetsov
#                              equation and converted internally)
#   n, b, A, BI                dimensionless
#   P(x)                       %
#
# The self-test at the bottom reproduces the Bårarp round 4 worked
# example of Ouchterlony (2005) within numerical tolerance. Reproducing
# a published example verifies the implementation of the equations; it
# does NOT validate the model for any particular site.
# ============================================================

from __future__ import annotations

import csv
import math
import os
import re
import warnings
from dataclasses import dataclass, field
from typing import Optional, Sequence

import numpy as np
import openpyxl

LN2 = math.log(2.0)

# Typical range of the Swebrec undulation parameter (b) reported in the
# literature. Values outside it are flagged with a warning, not rejected.
B_TYPICAL_MIN = 1.0
B_TYPICAL_MAX = 4.0


# ============================================================
# JOINT TERMS OF THE BLASTABILITY INDEX 
# ============================================================
_JCF_TABLE = {
    "tight": 1.0,
    "relaxed": 1.5,
    "gouge_filled": 2.0,
}

_JPA_VERSIONED = {
    "cunningham_1987": {
        "dip_out_of_face": 20,
        "strike_perpendicular_to_face": 30,
        "dip_into_face": 40,
    },
    "cunningham_2005": {
        "dip_out_of_face": 40,
        "strike_perpendicular_to_face": 30,
        "dip_into_face": 20,
    },
}


def _jpa_table_for_version(jpa_mapping_version: str) -> dict:
    """
    Return the JPA numeric table for the requested mapping version.

    Args:
        jpa_mapping_version: "cunningham_2005" (default baseline) or
            "cunningham_1987".

    Returns:
        dict: Case -> JPA rating.

    Raises:
        ValueError: If the version is not recognised.
    """
    try:
        return _JPA_VERSIONED[jpa_mapping_version]
    except KeyError:
        raise ValueError(
            f"unknown jpa_mapping_version {jpa_mapping_version!r}; "
            f"expected one of {sorted(_JPA_VERSIONED)}"
        )


def joint_condition_factor(joint_condition: str) -> float:
    """
    Return the joint condition factor JCF from a categorical description.

    Used only by the JF = JCF * JPS + JPA variant (later Cunningham
    variants). The baseline Cunningham (2005) KCO model uses JF = JPS + JPA
    and does not use JCF.

    Ratings:
        "tight"        -> JCF = 1.0
        "relaxed"      -> JCF = 1.5
        "gouge_filled" -> JCF = 2.0

    Args:
        joint_condition: One of "tight", "relaxed", "gouge_filled".

    Returns:
        float: Joint condition factor JCF (dimensionless).

    Raises:
        ValueError: If the category is not recognised.
    """
    try:
        return _JCF_TABLE[joint_condition]
    except KeyError:
        raise ValueError(
            f"unknown joint_condition {joint_condition!r}; "
            f"expected one of {sorted(_JCF_TABLE)}"
        )


def _joint_plane_angle_rating(orientation_case: str,
                              jpa_mapping_version: str = "cunningham_2005"
                              ) -> int:
    """
    Return the joint plane angle rating JPA from a categorical description.

    Two published mapping versions are supported:

        Cunningham 2005 (default):
            "dip_out_of_face"                -> 40
            "strike_perpendicular_to_face"   -> 30
            "dip_into_face"                  -> 20

        Cunningham 1987 (reversed):
            "dip_out_of_face"                -> 20
            "strike_perpendicular_to_face"   -> 30
            "dip_into_face"                  -> 40

    The correct version must be confirmed; the 2005 mapping is the
    active default because it matches the Cunningham 2005 rock-factor
    formulation used in this implementation. See :func:`_jpa_from_3d_planes`
    for the explicit 3-D helper that derives a category from joint
    (dip, dipdir) and face (dip, dipdir).

    Args:
        orientation_case: One of "dip_out_of_face",
            "strike_perpendicular_to_face", "dip_into_face".
        jpa_mapping_version: "cunningham_2005" (default) or
            "cunningham_1987".

    Returns:
        int: JPA rating (20, 30 or 40).

    Raises:
        ValueError: If the case or version is not recognised.
    """
    table = _jpa_table_for_version(jpa_mapping_version)
    try:
        return table[orientation_case]
    except KeyError:
        raise ValueError(
            f"unknown orientation_case {orientation_case!r}; "
            f"expected one of {sorted(table)}"
        )


def _plane3d_dip_vector(dip_deg: float, dipdir_deg: float) -> tuple[float, float, float]:
    """
    Return the unit vector pointing down the line of steepest descent of a
    plane with the given dip and dip direction.
    """
    d = math.radians(dip_deg)
    dd = math.radians(dipdir_deg)
    cos_d = math.cos(d)
    return (
        math.sin(dd) * cos_d,
        math.cos(dd) * cos_d,
        -math.sin(d),
    )


def _plane3d_outward_normal(dip_deg: float, dipdir_deg: float
                            ) -> tuple[float, float, float]:
    """
    Return the outward-pointing unit normal of a face with the given dip
    and dip direction.

    The face's rock mass lies on the upper side of the plane and the free
    space (excavation) lies on the lower side. The outward normal therefore
    points into the free space: it has the same horizontal trend as the face
    dip direction but a downward vertical component.
    """
    d = math.radians(dip_deg)
    dd = math.radians(dipdir_deg)
    sin_d = math.sin(d)
    return (
        math.sin(dd) * sin_d,
        math.cos(dd) * sin_d,
        -math.cos(d),
    )


def _angle_between_vectors(u: tuple, v: tuple) -> float:
    """Return the angle (degrees, 0-180) between two 3-D vectors."""
    dot = sum(ui * vi for ui, vi in zip(u, v))
    dot = max(-1.0, min(1.0, dot))
    return math.degrees(math.acos(dot))


def _jpa_from_3d_planes(joint_dip_deg: float,
                       joint_dipdir_deg: float,
                       face_dip_deg: float,
                       face_dipdir_deg: float,
                       subhorizontal_dip_deg: float = 30.0,
                       tolerance_deg: float = 45.0,
                       jpa_mapping_version: str = "cunningham_2005",
                       ) -> tuple[int, str, dict]:
    """
    Classify JPA by comparing the full 3-D joint and face planes.

    This is an explicit, inspectable geometric interpretation. The selected
    modelling assumptions are:

      * the joint is ``dip_out_of_face`` (least penalising) if its dip is
        below ``subhorizontal_dip_deg``, or if the joint's 3-D dip vector
        is within ``tolerance_deg`` of the face's 3-D outward normal;
      * the joint is ``dip_into_face`` if its dip vector is within
        ``tolerance_deg`` of the *opposite* of the face's outward normal;
      * otherwise it is treated as ``strike_perpendicular_to_face``.

    Geometric quantities returned:
      - ``alpha_deg``: the 3-D angle between the joint's dip vector and the
        face's outward normal (0-180°).
      - ``tolerance_deg``: the same threshold used for the classification.
      - ``subhorizontal_dip_deg``: the flat-joint override threshold.

    Args:
        joint_dip_deg: Dip of the joint set (0-90°).
        joint_dipdir_deg: Dip direction of the joint set (0-360°).
        face_dip_deg: Dip of the free face (0-90°).
        face_dipdir_deg: Dip direction of the free face (0-360°).
        subhorizontal_dip_deg: Dip below which a set is treated as sub-
            horizontal (default 30°).
        tolerance_deg: Angular half-width of the "out of face" and
            "into face" cones (default 45°).
        jpa_mapping_version: "cunningham_2005" (default; 40/30/20) or
            "cunningham_1987".

    Returns:
        tuple[int, str, dict]: (JPA rating, case string, geometry dict).
    """
    table = _jpa_table_for_version(jpa_mapping_version)

    # Flat-lying joints are explicitly treated as the least penalising case.
    if joint_dip_deg < subhorizontal_dip_deg:
        case = "dip_out_of_face"
        jpa = table[case]
        geometry = {
            "subhorizontal_dip_deg": subhorizontal_dip_deg,
            "joint_dip_deg": joint_dip_deg,
            "classification_basis": "sub-horizontal override",
        }
        return jpa, case, geometry

    d_joint = _plane3d_dip_vector(joint_dip_deg, joint_dipdir_deg)
    n_face = _plane3d_outward_normal(face_dip_deg, face_dipdir_deg)
    alpha = _angle_between_vectors(d_joint, n_face)

    if alpha < tolerance_deg:
        case = "dip_out_of_face"
    elif alpha > 180.0 - tolerance_deg:
        case = "dip_into_face"
    else:
        case = "strike_perpendicular_to_face"
    jpa = table[case]

    geometry = {
        "alpha_deg": alpha,
        "tolerance_deg": tolerance_deg,
        "subhorizontal_dip_deg": subhorizontal_dip_deg,
        "joint_dip_vector": d_joint,
        "face_outward_normal": n_face,
        "classification_basis": "3-D dip-vector vs face-outward-normal",
    }
    return jpa, case, geometry


def jps_from_joint_spacing(mean_joint_spacing_m: float,
                           burden_m: float,
                           spacing_m: float) -> int:
    """
    Return the joint plane spacing rating JPS.

    Classification as it appears in the KCO literature (Cunningham 1987
    variant with the drilling-pattern reference length sqrt(B*S)):

        JPS = 10   if Sj < 0.1 m
        JPS = 20   if 0.1 m <= Sj < 0.3 m
        JPS = 50   if 0.3 m <= Sj < 0.95*sqrt(B*S)
        JPS = 80   if Sj >= 0.95*sqrt(B*S)

    Boundary handling (explicit, documented choices):
      - all interval boundaries are closed on the left (>=) and open on
        the right (<), so every spacing maps to exactly one rating;
      - the published texts leave the band between 0.95*sqrt(B*S) and
        sqrt(B*S) ambiguous ("> oversize" vs ">= 0.95 sqrt(BS)"). Here
        the 0.95*sqrt(B*S) threshold is used as the lower edge of the
        JPS = 80 class, which removes the gap. This is a selected
        modelling assumption;
      - if 0.95*sqrt(B*S) <= 0.3 m (a very tight drilling pattern), the
        JPS = 50 class is empty and spacings >= 0.3 m rate 80.

    Args:
        mean_joint_spacing_m: Mean joint spacing Sj (m). In this project
            it comes from the SPACING analysis outputs.
        burden_m: Burden B (m), used in the pattern reference length.
        spacing_m: Hole spacing S (m), used in the pattern reference
            length.

    Returns:
        int: JPS rating (10, 20, 50 or 80).

    Raises:
        ValueError: If any argument is not strictly positive.
    """
    if mean_joint_spacing_m <= 0:
        raise ValueError("mean_joint_spacing_m must be > 0")
    if burden_m <= 0 or spacing_m <= 0:
        raise ValueError("burden_m and spacing_m must be > 0")

    pattern_length = math.sqrt(burden_m * spacing_m)
    upper = 0.95 * pattern_length

    if mean_joint_spacing_m < 0.1:
        return 10
    if mean_joint_spacing_m < 0.3:
        return 20
    if mean_joint_spacing_m < upper:
        return 50
    return 80


def joint_factor(jps: float,
                 jpa: float,
                 jcf: Optional[float] = None) -> float:
    """
    Return the joint factor JF.

    Two published forms are supported:

        JF = JPS + JPA              (Cunningham 1987, baseline;
                                     used when jcf is None)
        JF = JCF * JPS + JPA        (later variants; used when a joint
                                     condition factor is supplied)

    Args:
        jps: Joint plane spacing rating, see :func:`jps_from_joint_spacing`.
        jpa: Joint plane angle rating, see :func:`_joint_plane_angle_rating`.
        jcf: Optional joint condition factor, see
            :func:`joint_condition_factor`. None means the baseline
            formulation without JCF.

    Returns:
        float: Joint factor JF (dimensionless).
    """
    if jcf is None:
        return jps + jpa
    return jcf * jps + jpa


# ============================================================
# DENSITY AND HARDNESS TERMS
# ============================================================
def rock_density_influence(rho_kg_m3: float) -> float:
    """
    Return the rock density influence RDI.

    Lilly (1986) / Cunningham (1987), as quoted in Ouchterlony (2005)
    eq. (11e):

        RDI = 0.025 * rho - 50      with rho in kg/m3

    (equivalently RDI = 25*rho - 50 with rho in t/m3).

    Args:
        rho_kg_m3: Intact rock density (kg/m3). Must be > 0.

    Returns:
        float: RDI term (dimensionless).

    Raises:
        ValueError: If the density is not strictly positive.
    """
    if rho_kg_m3 <= 0:
        raise ValueError("rho_kg_m3 must be > 0")
    return 0.025 * rho_kg_m3 - 50.0


def hardness_factor(youngs_modulus_gpa: float, ucs_mpa: float) -> float:
    """
    Return the hardness factor HF.

    Lilly (1986) / Cunningham (1987), as quoted in Ouchterlony (2005)
    eq. (11e) and Ouchterlony & Sanchidrian (2019) eq. (21):

        HF = E / 3          if E < 50 GPa
        HF = UCS / 5        otherwise

    Units: E in GPa, UCS (sigma_c) in MPa.

    Args:
        youngs_modulus_gpa: Young's modulus E (GPa). Must be > 0.
        ucs_mpa: Uniaxial compressive strength sigma_c (MPa). Must be > 0.

    Returns:
        float: HF term (dimensionless).

    Raises:
        ValueError: If either input is not strictly positive.
    """
    if youngs_modulus_gpa <= 0:
        raise ValueError("youngs_modulus_gpa must be > 0")
    if ucs_mpa <= 0:
        raise ValueError("ucs_mpa must be > 0")
    if youngs_modulus_gpa < 50.0:
        return youngs_modulus_gpa / 3.0
    return ucs_mpa / 5.0


# ============================================================
# BLASTABILITY INDEX AND ROCK FACTOR
# ============================================================
def rock_mass_description(rock_mass_case: str,
                          jf: Optional[float] = None) -> float:
    """
    Return the Rock Mass Description rating RMD.

    Ouchterlony (2005), Eq. (11e):

        powdery/friable rock mass  -> RMD = 10
        jointed rock mass          -> RMD = JF
        massive rock               -> RMD = 50

    For the jointed case, JF must already have been calculated from
    JPS/JPA (and JCF only if that formulation is deliberately enabled).

    Args:
        rock_mass_case: "powdery_friable", "jointed", or "massive".
        jf: Joint factor, required when ``rock_mass_case == "jointed"``.

    Returns:
        float: RMD rating.

    Raises:
        ValueError: For an unknown case or missing JF in the jointed case.
    """
    if rock_mass_case == "powdery_friable":
        return 10.0

    if rock_mass_case == "jointed":
        if jf is None:
            raise ValueError(
                "rock_mass_case='jointed' requires a calculated JF"
            )
        return float(jf)

    if rock_mass_case == "massive":
        return 50.0

    raise ValueError(
        "rock_mass_case must be 'powdery_friable', "
        "'jointed', or 'massive'"
    )


def blastability_index(rmd: float,
                       rdi: float,
                       hf: float) -> float:
    """
    Return the Cunningham (2005) blastability index BI.

    The independent JF term is excluded; JF already enters through the
    jointed-rock RMD when that rock-mass case is selected:

        BI = RMD + RDI + HF

    where RMD is computed by :func:`rock_mass_description` from the
    selected rock-mass case (RMD = 10 for powdery/friable, RMD = JF
    for jointed, RMD = 50 for massive).

    Args:
        rmd: Rock mass description rating.
        rdi: Rock density influence, see :func:`rock_density_influence`.
        hf: Hardness factor, see :func:`hardness_factor`.

    Returns:
        float: Blastability index BI (dimensionless).
    """
    return rmd + rdi + hf


def rock_factor_A(bi: float) -> float:
    """
    Return the Kuz-Ram / KCO rock factor A from the blastability index.

        A = 0.06 * BI

    Uses the Cunningham (2005) blastability index BI = RMD + RDI + HF.
    Reported values of A span roughly 1.7 to 21 (2019 review), against
    the narrower 7 to 13 of the original Kuznetsov ratings.

    Args:
        bi: Blastability index, see :func:`blastability_index`.

    Returns:
        float: Rock factor A (dimensionless).
    """
    return 0.06 * bi


# ============================================================
# POWDER FACTOR
# ============================================================
def calculate_powder_factor(charge_per_hole_kg: float,
                            burden_m: float,
                            spacing_m: float,
                            bench_height_m: float) -> float:
    """
    Return the powder factor q computed from the nominal breakage volume.

        V0 = B * S * H              (m3 per hole)
        q  = Q / V0                 (kg/m3)

    H is the bench height (m). Subdrill is ignored for the powder-factor
    volume; the total charge above grade is divided by the in-situ volume
    defined by the bench geometry.

    Args:
        charge_per_hole_kg: Charge per hole Q (kg).
        burden_m: Burden B (m).
        spacing_m: Spacing S (m).
        bench_height_m: Bench height H (m).

    Returns:
        float: Powder factor q (kg/m3).

    Raises:
        ValueError: If any argument is not strictly positive.
    """
    if charge_per_hole_kg <= 0:
        raise ValueError("charge_per_hole_kg must be > 0")
    if burden_m <= 0 or spacing_m <= 0 or bench_height_m <= 0:
        raise ValueError("burden_m, spacing_m and bench_height_m "
                         "must be > 0")
    return charge_per_hole_kg / (burden_m * spacing_m * bench_height_m)


# ============================================================
# UNIFORMITY INDEX n — Cunningham (2005), Ouchterlony & Sanchidrian
# (2019) Eq. (48)
# ============================================================
def uniformity_index_n(
        burden_m: float,
        spacing_m: float,
        hole_diameter_mm: float,
        drill_accuracy_sd_m: float,
        charge_length_m: float,
        bench_height_m: float,
        rock_factor_A: float,
        timing_scatter_factor_ns: float) -> float:
    """
    Return the Cunningham (2005) uniformity index n.

    Corrected/confirmed form from Ouchterlony & Sanchidrian (2019)
    eq. (48), with the (A/6)^0.3 term included:

        n = n_s * sqrt(2 - 30*B/d) * sqrt((1 + S/B)/2)
            * (1 - W/B) * (L/H)^0.3 * (A/6)^0.3

    where C(n) = 1 (no site-specific calibration).

    Args:
        burden_m: Burden B (m). Must be > 0.
        spacing_m: Spacing S (m). Must be > 0.
        hole_diameter_mm: Hole diameter d (mm). Must be > 0.
        drill_accuracy_sd_m: Drilling accuracy standard deviation W (m).
            Must satisfy 0 <= W < B.
        charge_length_m: Charge length L (m), above grade.
        bench_height_m: Bench height H (m). Must be > 0.
        rock_factor_A: Rock factor A, see :func:`rock_factor_A`.
            Must be > 0.
        timing_scatter_factor_ns: Timing-scatter factor n_s. Must be > 0.
            If timing-scatter data are not available this must be
            supplied explicitly; do not invent a value.

    Returns:
        float: Uniformity index n.

    Raises:
        ValueError: If any validation fails, or if the resulting n is
            not strictly positive.
    """
    if burden_m <= 0:
        raise ValueError("burden_m must be > 0")
    if spacing_m <= 0:
        raise ValueError("spacing_m must be > 0")
    if hole_diameter_mm <= 0:
        raise ValueError("hole_diameter_mm must be > 0")
    if bench_height_m <= 0:
        raise ValueError("bench_height_m must be > 0")
    if charge_length_m <= 0:
        raise ValueError("charge_length_m must be > 0")
    if rock_factor_A <= 0:
        raise ValueError("rock_factor_A must be > 0")
    if timing_scatter_factor_ns is None:
        raise ValueError(
            "timing_scatter_factor_ns is required; set it explicitly "
            "or compute it from the in-row delay scatter."
        )
    if timing_scatter_factor_ns <= 0:
        raise ValueError("timing_scatter_factor_ns must be > 0")
    if not (0.0 <= drill_accuracy_sd_m < burden_m):
        raise ValueError("drill_accuracy_sd_m must satisfy 0 <= W < B")

    geometry_inner = 2.0 - 30.0 * burden_m / hole_diameter_mm
    if geometry_inner <= 0:
        raise ValueError(
            "Cunningham 2005 geometry term is non-positive."
        )

    term_geometry = math.sqrt(geometry_inner)
    term_spacing = math.sqrt(
        (1.0 + spacing_m / burden_m) / 2.0
    )
    term_accuracy = 1.0 - drill_accuracy_sd_m / burden_m
    term_length = (charge_length_m / bench_height_m) ** 0.3
    term_rock = (rock_factor_A / 6.0) ** 0.3

    n = (
        timing_scatter_factor_ns
        * term_geometry
        * term_spacing
        * term_accuracy
        * term_length
        * term_rock
    )

    if n <= 0:
        raise ValueError("uniformity index n must be > 0")

    return n


# ============================================================
# SHIFT FACTOR g(n) — OPTIONAL, later-generation correction
# ============================================================
def shift_factor_g(n: float, mode: str = "no_shift") -> float:
    """
    Return g(n) for the X50 equation.

    Ouchterlony (2005) Eq. (11b) allows two interpretations of the
    median size X50: the uncorrected (g(n) = 1) form used by the
    Swebrec/KCO baseline, and the mean-to-median shift of the Rosin-
    Rammler distribution.

    Modes:
        "no_shift":
            g(n) = 1
            Baseline. Ouchterlony (2005) explicitly accepts this because
            it was uncertain whether the shift factor was needed with the
            Swebrec function.
        "mean_to_median_shift":
            g(n) = (ln 2)^(1/n) / Gamma(1 + 1/n)   < 1
            Theoretical median-over-mean ratio of the Rosin-Rammler
            distribution. This is a sensitivity option, not the baseline.

    Literature/transcription note: for the Bårarp example (n = 1.17) the
    2005 paper prints g(n) = 0.659 while direct evaluation of the
    formula above gives 0.773. This module returns the value of the
    formula; the discrepancy with the printed number is flagged as a
    probable transcription issue in the source, not resolved here.

    Args:
        n: Cunningham uniformity index. Must be > 0.
        mode: "no_shift" or "mean_to_median_shift". Defaults to
            "no_shift".

    Returns:
        float: Shift factor g(n).

    Raises:
        ValueError: If n is not strictly positive, or if ``mode`` is not
            one of the two supported choices.
    """
    if n <= 0:
        raise ValueError("n must be > 0")
    if mode == "no_shift":
        return 1.0
    if mode == "mean_to_median_shift":
        return LN2 ** (1.0 / n) / math.gamma(1.0 + 1.0 / n)
    raise ValueError(
        "shift mode must be 'no_shift' or 'mean_to_median_shift'"
    )


# ============================================================
# MEDIAN SIZE X50 — Kuznetsov / KCO, eq. (11b)
# ============================================================
def x50_kuznetsov(rock_factor_a: float,
                  charge_per_hole_kg: float,
                  powder_factor_kg_m3: float,
                  s_anfo_pct: float,
                  g_n: float = 1.0,
                  timing_factor: float = 1.0) -> float:
    """
    Return the median (50 % passing) fragment size X50 in CENTIMETRES.

    Ouchterlony (2005) eq. (11b), identical to Cunningham (1987) as
    written in Ouchterlony & Sanchidrian (2019) eq. (24):

        X50 = g(n) * A * Q^(1/6) * q^(-0.8) * (115 / s_ANFO)^(19/30)

    Baseline: g(n) = 1 and timing_factor = 1. The shift factor g(n) and
    the Cunningham (2005) timing factor A_T are later-generation /
    sensitivity options. The caller is responsible for selecting and
    passing the desired g(n); :func:`predict_kco` does this from
    ``design.shift_factor_mode``.

    The caller is also responsible for converting to mm for final outputs
    (:func:`predict_kco` does this).

    Args:
        rock_factor_a: Rock factor A. Must be > 0.
        charge_per_hole_kg: Charge per hole Q (kg). Must be > 0.
        powder_factor_kg_m3: Powder factor q (kg/m3). Must be > 0.
        s_anfo_pct: Explosive weight strength relative to ANFO (%).
            ANFO itself is 100. Must be > 0.
        g_n: Shift factor g(n) applied to the median size. Must be > 0.
            Defaults to 1.0 (no shift, the Swebrec/KCO baseline).
        timing_factor: Cunningham (2005) timing factor A_T. Defaults
            to 1.0 (baseline).

    Returns:
        float: Median fragment size X50 (cm).

    Raises:
        ValueError: If any input is non-positive.
    """
    if rock_factor_a <= 0:
        raise ValueError("rock_factor_a must be > 0")
    if charge_per_hole_kg <= 0:
        raise ValueError("charge_per_hole_kg must be > 0")
    if powder_factor_kg_m3 <= 0:
        raise ValueError("powder_factor_kg_m3 must be > 0")
    if s_anfo_pct <= 0:
        raise ValueError("s_anfo_pct must be > 0")
    if g_n <= 0:
        raise ValueError("g_n must be > 0")

    return (g_n * timing_factor * rock_factor_a
            * charge_per_hole_kg ** (1.0 / 6.0)
            * powder_factor_kg_m3 ** (-0.8)
            * (115.0 / s_anfo_pct) ** (19.0 / 30.0))


# ============================================================
# IN-SITU BLOCK SIZE FROM THE DFN BLOCK-VOLUME DISTRIBUTION
# ============================================================
def block_volume_to_equivalent_cube(volume_m3):
    """
    Convert block volume(s) to the edge length of a cube of equal volume.

        D_eq,cube = V^(1/3)

    This is one of several possible volume-to-length conversions. Which
    one represents the "in-situ block size" of the KCO xmax rule is a
    user-configurable interpretation, not a uniquely defined published
    conversion.

    Args:
        volume_m3: Block volume(s) (m3), scalar or array.

    Returns:
        Equivalent cube edge length(s) (m), same shape as input.
    """
    return np.asarray(volume_m3, dtype=float) ** (1.0 / 3.0)


def block_volume_to_equivalent_sphere(volume_m3):
    """
    Convert block volume(s) to the diameter of a sphere of equal volume.

        D_eq,sphere = (6 V / pi)^(1/3)

    See the note in :func:`block_volume_to_equivalent_cube`: the choice
    of conversion is a user-configurable interpretation.

    Args:
        volume_m3: Block volume(s) (m3), scalar or array.

    Returns:
        Equivalent sphere diameter(s) (m), same shape as input.
    """
    return (6.0 * np.asarray(volume_m3, dtype=float) / math.pi) ** (1.0 / 3.0)


_BLOCK_STATISTIC_PERCENTILE = {
    "median": 50.0,
    "p80": 80.0,
    "p95": 95.0,
    "maximum": 100.0,
}


def characteristic_block_size_from_distribution(
        block_volumes_m3,
        block_size_method: str,
        block_statistic: Optional[str] = None,
        block_percentile: Optional[float] = None,
) -> tuple[float, dict]:
    """
    Derive one characteristic in-situ block size from a DFN distribution.

    The DFN workflow produces a distribution of block volumes; the KCO
    xmax rule needs a single length. Neither the volume-to-length
    conversion nor the representative statistic is uniquely defined by
    the KCO sources, so BOTH must be chosen explicitly by the user.
    The choices made are returned so they can be reported.

    Args:
        block_volumes_m3: Iterable of block volumes (m3). Non-finite and
            non-positive entries are discarded.
        block_size_method: "equivalent_cube" (D = V^(1/3)) or
            "equivalent_sphere" (D = (6V/pi)^(1/3)). No default: the
            choice is a modelling assumption the user must make.
        block_statistic: One of "median", "p80", "p95", "maximum";
            mutually exclusive with ``block_percentile``.
        block_percentile: Percentile of the volume distribution to use
            (0-100]; mutually exclusive with ``block_statistic``.

    Returns:
        tuple[float, dict]: (characteristic size in m, info dict). The
        info dict records "block_size_method", "block_statistic",
        "block_percentile", "block_volume_m3" (the selected volume),
        "block_size_value_m", "n_blocks" and "warnings" (a list;
        non-empty when the maximum statistic is chosen, since the
        absolute maximum is sensitive to model domain boundaries and
        outliers).

    Raises:
        ValueError: If the method is not recognised, if neither or both
            of the statistic/percentile selectors are given, or if no
            valid volumes remain after filtering.
    """
    if block_size_method == "equivalent_cube":
        convert = block_volume_to_equivalent_cube
    elif block_size_method == "equivalent_sphere":
        convert = block_volume_to_equivalent_sphere
    else:
        raise ValueError(
            f"unknown block_size_method {block_size_method!r}; expected "
            "'equivalent_cube' or 'equivalent_sphere' (or use "
            "'user_defined' with an explicit in_situ_block_size_m in the "
            "BlastDesign)"
        )

    if (block_statistic is None) == (block_percentile is None):
        raise ValueError("supply exactly one of block_statistic or "
                         "block_percentile")
    if block_statistic is not None:
        if block_statistic not in _BLOCK_STATISTIC_PERCENTILE:
            raise ValueError(
                f"unknown block_statistic {block_statistic!r}; expected "
                f"one of {sorted(_BLOCK_STATISTIC_PERCENTILE)}"
            )
        percentile = _BLOCK_STATISTIC_PERCENTILE[block_statistic]
    else:
        if not 0.0 < block_percentile <= 100.0:
            raise ValueError("block_percentile must be in (0, 100]")
        percentile = float(block_percentile)

    v = np.asarray(list(block_volumes_m3), dtype=float).ravel()
    v = v[np.isfinite(v) & (v > 0.0)]
    if v.size == 0:
        raise ValueError("no valid (finite, positive) block volumes supplied")

    volume = float(np.percentile(v, percentile))
    size_m = float(convert(volume))

    warnings: list[str] = []
    if percentile >= 100.0:
        warnings.append(
            "in-situ block size is based on the distribution maximum, "
            "which is sensitive to DFN domain boundaries and outliers; "
            "consider a high percentile (e.g. p95) instead"
        )

    info = {
        "block_size_method": block_size_method,
        "block_statistic": block_statistic,
        "block_percentile": percentile,
        "block_volume_m3": volume,
        "block_size_value_m": size_m,
        "n_blocks": int(v.size),
        "warnings": warnings,
    }
    return size_m, info


# ============================================================
# MAXIMUM SIZE Xmax — KCO eq. (11d)
# ============================================================
def xmax_kco(in_situ_block_size_m: float,
             burden_m: float,
             spacing_m: float) -> tuple[float, str]:
    """
    Return the maximum fragment size Xmax (m) and what controls it.

    Ouchterlony (2005) eq. (11d):

        Xmax = min(in-situ block size, S, B)

    The physical argument in the source is that a fragment can be no
    larger than either the natural block delimited by the joint network
    or the slab bounded by the drilling pattern. The paper itself labels
    the expression "tentative".

    Args:
        in_situ_block_size_m: Characteristic in-situ block size (m),
            derived by an explicit, user-selected method (see
            :func:`characteristic_block_size_from_distribution`).
        burden_m: Burden B (m).
        spacing_m: Spacing S (m).

    Returns:
        tuple[float, str]: (Xmax in m, governing quantity, one of
        "in-situ block", "burden", "spacing").

    Raises:
        ValueError: If any argument is not strictly positive.
    """
    if in_situ_block_size_m <= 0 or burden_m <= 0 or spacing_m <= 0:
        raise ValueError("all xmax candidates must be > 0")
    candidates = {
        "in-situ block": in_situ_block_size_m,
        "burden": burden_m,
        "spacing": spacing_m,
    }
    governed_by = min(candidates, key=candidates.get)
    return candidates[governed_by], governed_by


# ============================================================
# UNDULATION PARAMETER b
# ============================================================
def b_parameter(xmax: float, x50: float, n: float) -> float:
    """
    Return the Swebrec undulation parameter b.

    Ouchterlony (2005), first line of eq. (11c), obtained by equating
    the Swebrec and Rosin-Rammler slopes at X50 (eq. 3 inverted):

        b = 2 * ln(2) * n * ln(Xmax / X50)

    Args:
        xmax: Maximum fragment size, any length unit.
        x50: Median fragment size, in the SAME unit as ``xmax``.
        n: Cunningham uniformity index. Must be > 0.

    Returns:
        float: Undulation parameter b (dimensionless). Values typically
            fall in the range 1-4; values outside that range may still
            occur and are flagged by :func:`predict_kco` with a warning
            rather than rejected.

    Raises:
        ValueError: If Xmax > X50 > 0 does not hold, if n <= 0, or if
            the computed b is not strictly positive.
    """
    if not xmax > x50 > 0:
        raise ValueError(f"require Xmax > X50 > 0, got xmax={xmax}, "
                         f"x50={x50}")
    if n <= 0:
        raise ValueError("n must be > 0")
    b = 2.0 * LN2 * n * math.log(xmax / x50)
    if b <= 0:
        raise ValueError(f"computed b = {b} is not positive; check inputs")
    return b


# ============================================================
# SWEBREC DISTRIBUTION — Ouchterlony (2005) eq. (11a)
# ============================================================
def swebrec_passing(x, x50: float, xmax: float, b: float) -> np.ndarray:
    """
    Return the Swebrec cumulative percentage passing P(x), in percent.

    Ouchterlony (2005) eq. (11a):

        P(x) = 100 / (1 + [ ln(Xmax/x) / ln(Xmax/X50) ]^b )
                                            for 0 < x < Xmax
        P(x <= 0)    = 0
        P(x >= Xmax) = 100

    By construction P(X50) = 50 % and P(Xmax) = 100 %. Unlike
    Rosin-Rammler, the distribution has a finite upper limit and a
    logarithmic (not power-law) fines asymptote.

    Args:
        x: Fragment size(s), same length unit as ``x50`` and ``xmax``.
        x50: Median size (50 % passing).
        xmax: Maximum size (100 % passing).
        b: Undulation parameter, see :func:`b_parameter`.

    Returns:
        np.ndarray: Percentage passing in [0, 100].

    Raises:
        ValueError: If Xmax > X50 > 0 does not hold or b <= 0.
    """
    if not xmax > x50 > 0:
        raise ValueError(f"require Xmax > X50 > 0, got xmax={xmax}, "
                         f"x50={x50}")
    if b <= 0:
        raise ValueError("b must be > 0")

    x = np.asarray(x, dtype=float)
    out = np.zeros_like(x)
    inside = (x > 0.0) & (x < xmax)
    ratio = np.log(xmax / x[inside]) / math.log(xmax / x50)
    out[inside] = 100.0 / (1.0 + ratio ** b)
    out[x >= xmax] = 100.0
    return out


def swebrec_size_at_passing(passing_pct, x50: float, xmax: float,
                            b: float) -> np.ndarray:
    """
    Invert the Swebrec function: return the size X_P at a given % passing.

        X_P = Xmax * exp( -ln(Xmax/X50) * (100/P - 1)^(1/b) )

    Args:
        passing_pct: Percentage passing P in (0, 100].
        x50: Median size.
        xmax: Maximum size, same unit.
        b: Undulation parameter.

    Returns:
        np.ndarray: Fragment size(s) X_P, same unit as ``x50``. NaN for
            P outside (0, 100].
    """
    p = np.asarray(passing_pct, dtype=float)
    out = np.full_like(p, np.nan)
    valid = (p > 0.0) & (p <= 100.0)
    ratio = (100.0 / p[valid] - 1.0) ** (1.0 / b)
    out[valid] = xmax * np.exp(-math.log(xmax / x50) * ratio)
    return out


# ============================================================
# ROSIN-RAMMLER — REFERENCE / COMPARISON MODEL ONLY
# ============================================================
def rosin_rammler_passing(x, x50: float, n: float) -> np.ndarray:
    """
    Return the Rosin-Rammler cumulative percentage passing.

    REFERENCE MODEL ONLY (reference_model_only = True): this is the
    distribution of the original Kuz-Ram model, kept exclusively so the
    KCO/Swebrec prediction can be compared against it. It must not be
    used as the final post-blast prediction, because it underestimates
    fines and has no upper size limit.

        P_RR(x) = 100 * [1 - 2^(-(x/X50)^n)]

    Args:
        x: Fragment size(s), same unit as ``x50``.
        x50: Median size (50 % passing).
        n: Cunningham uniformity index.

    Returns:
        np.ndarray: Percentage passing in [0, 100].
    """
    x = np.asarray(x, dtype=float)
    out = np.zeros_like(x)
    pos = x > 0.0
    out[pos] = 100.0 * (1.0 - 2.0 ** (-((x[pos] / x50) ** n)))
    return out


# ============================================================
# INPUT CONTAINER
# ============================================================
@dataclass
class BlastDesign:
    """
    Complete set of inputs for one KCO prediction.

    Units (fixed, documented, and never silently converted):

        burden_m, spacing_m, bench_height_m          m
        hole_diameter_mm                             mm
        drill_accuracy_sd_m                          m   (0 <= W < B)
        bottom_charge_m, column_charge_m             m
        total_charge_m (above grade)                 m
        subdrill_m                                   m
        charge_per_hole_kg                           kg/hole
        powder_factor_reported_kg_m3                 kg/m3
        s_anfo_pct                                   %
        rock_density_kg_m3                           kg/m3
        ucs_mpa                                      MPa
        youngs_modulus_gpa                           GPa
        mean_joint_spacing_m, in_situ_block_size_m   m

    Model-generation switches (baseline = Cunningham 2005 KCO):
        timing_factor and shift_factor_mode are later-generation / sensitivity
        choices. The baseline uses shift_factor_mode = "no_shift" (g(n)=1); the
        "mean_to_median_shift" option must be enabled deliberately.

    Methodological choices the user must make explicitly (scientifically
    neutral defaults are provided only where a baseline is defined by
    the cited source):
        rock_mass_case, joint_condition (only if jf_includes_jcf),
        jpa_case or jpa_rating, powder_factor_mode,
        block_size_method + block_statistic / block_percentile.

    Attributes:
        name: Label for the round or site, used in outputs.
        hole_diameter_mm: Hole diameter D (mm).
        burden_m: Burden B (m).
        spacing_m: Spacing S (m).
        bench_height_m: Bench height H (m).
        total_charge_m: Total charge length above grade Ltot (m).
        bottom_charge_m: Bottom charge length Lb (m).
        column_charge_m: Column charge length Lc (m).
        subdrill_m: Subdrill (m); recorded for reporting only.
        drill_accuracy_sd_m: Drilling accuracy standard deviation W (m).
        charge_per_hole_kg: Charge per hole Q (kg).
        powder_factor_reported_kg_m3: Powder factor q (kg/m3) as reported
            by the operation; None if unknown.
        powder_factor_mode: "reported" uses the reported value;
            "calculated" uses Q/(B*S*H) where H is the bench height. When
            both are available the other is still computed for cross-checking.
        q_tolerance_pct: Warn when reported and calculated powder factors
            differ by more than this percentage. Default 10.
        s_anfo_pct: Explosive weight strength relative to ANFO (%).
        explosive_name: Product name, for reporting.
        rock_density_kg_m3: Intact rock density rho (kg/m3).
        ucs_mpa: Uniaxial compressive strength sigma_c (MPa).
        youngs_modulus_gpa: Young's modulus E (GPa).
        rock_mass_case: Cunningham (2005) rock-mass category:
            "powdery_friable" (RMD = 10), "jointed" (RMD = JF) or
            "massive" (RMD = 50). RMD is computed by
            :func:`rock_mass_description` from this case and the JF.
        jpa_mapping_version: JPA numeric mapping version. Either
            "cunningham_2005" (baseline; 40/30/20) or "cunningham_1987"
            (20/30/40). Must be confirmed with the supervisors; defaults
            to the 2005 mapping that matches the Cunningham 2005 rock-factor
            formulation.
        jpa_tolerance_deg: Half-width (degrees) of the 3-D "out of face"
            and "into face" cones around the face outward normal. Default
            45.0°, which is a modelling assumption to confirm with the
            supervisors.
        jpa_case: Joint plane angle category ("dip_out_of_face",
            "strike_perpendicular_to_face", "dip_into_face"); mutually
            exclusive with jpa_rating.
        jpa_rating: Explicit numeric JPA (20/30/40); overrides jpa_case
            when set. If provided, it is used exactly as given regardless
            of jpa_mapping_version.
        jf_includes_jcf: Use JF = JCF*JPS + JPA (later variant) instead
            of the baseline JF = JPS + JPA.
        joint_condition: "tight" (1.0), "relaxed" (1.5) or
            "gouge_filled" (2.0); only used when jf_includes_jcf is True.
        mean_joint_spacing_m: Mean joint spacing Sj (m), from the SPACING
            analysis in this project.
        block_size_method: "equivalent_cube", "equivalent_sphere" or
            "user_defined". With "user_defined",
            ``in_situ_block_size_m`` is used directly and no DFN
            conversion is applied.
        block_statistic: "median", "p80", "p95" or "maximum"; used with
            a DFN block-volume distribution. Mutually exclusive with
            block_percentile.
        block_percentile: Percentile (0-100] of the block-volume
            distribution; mutually exclusive with block_statistic.
        in_situ_block_size_m: Characteristic in-situ block size (m);
            required when block_size_method = "user_defined".
        timing_factor: Cunningham (2005) timing factor A_T. Baseline 1.0.
        timing_scatter_factor_ns: In-row delay scatter factor n_s
            (Cunningham 2005). Must be > 0. If the delay-scatter data
            needed to compute it are not available, leave as None and the
            workflow will raise a clear error.
        shift_factor_mode: How g(n) is treated in the X50 equation.
            Either "no_shift" (g(n)=1, baseline) or
            "mean_to_median_shift" (g(n)=(ln2)^(1/n)/Gamma(1+1/n)).
        charge_sum_tolerance_pct: Warn when Lb + Lc differs from Ltot by
            more than this percentage. Default 5.
    """
    name: str = "VARENNE"

    # --- geometry (m / mm) ---
    hole_diameter_mm: float = 0.0
    burden_m: float = 0.0
    spacing_m: float = 0.0
    bench_height_m: float = 0.0
    total_charge_m: float = 0.0
    bottom_charge_m: float = 0.0
    column_charge_m: float = 0.0
    subdrill_m: float = 0.0
    drill_accuracy_sd_m: float = 0.0

    # --- explosive ---
    charge_per_hole_kg: float = 0.0
    powder_factor_reported_kg_m3: Optional[float] = None
    powder_factor_mode: str = "reported"
    q_tolerance_pct: float = 10.0
    s_anfo_pct: float = 100.0
    explosive_name: str = "ANFO"

    # --- rock mass ---
    rock_density_kg_m3: float = 0.0
    ucs_mpa: float = 0.0
    youngs_modulus_gpa: float = 0.0

    # --- rock factor formulation ---
    rock_mass_case: str = "jointed"
    jpa_mapping_version: str = "cunningham_2005"
    jpa_tolerance_deg: float = 45.0
    jpa_case: Optional[str] = None
    jpa_rating: Optional[int] = None
    jf_includes_jcf: bool = False
    joint_condition: Optional[str] = None
    mean_joint_spacing_m: float = 0.0

    # --- in-situ block size (DFN) ---
    block_size_method: str = "user_defined"
    block_statistic: Optional[str] = None
    block_percentile: Optional[float] = None
    in_situ_block_size_m: Optional[float] = None

    # --- later-generation / sensitivity options (baseline: neutral) ---
    timing_factor: float = 1.0
    timing_scatter_factor_ns: Optional[float] = None
    shift_factor_mode: str = "no_shift"

    charge_sum_tolerance_pct: float = 5.0


def _validate_design(d: BlastDesign) -> list[str]:
    """
    Validate physical inputs of a BlastDesign; return non-fatal warnings.

    Raises ValueError for physically or mathematically invalid inputs;
    returns a list of warning strings for unusual but not necessarily
    invalid combinations (e.g. Lb + Lc != Ltot beyond tolerance).

    Args:
        d: The design to validate.

    Returns:
        list[str]: Warnings collected during validation.

    Raises:
        ValueError: On any invalid input.
    """
    checks = [
        ("burden_m", d.burden_m),
        ("spacing_m", d.spacing_m),
        ("bench_height_m", d.bench_height_m),
        ("hole_diameter_mm", d.hole_diameter_mm),
        ("charge_per_hole_kg", d.charge_per_hole_kg),
        ("s_anfo_pct", d.s_anfo_pct),
        ("rock_density_kg_m3", d.rock_density_kg_m3),
        ("youngs_modulus_gpa", d.youngs_modulus_gpa),
        ("ucs_mpa", d.ucs_mpa),
        ("total_charge_m", d.total_charge_m),
    ]
    for label, value in checks:
        if value <= 0:
            raise ValueError(f"{label} must be > 0 (got {value})")
    if d.bottom_charge_m < 0 or d.column_charge_m < 0:
        raise ValueError("bottom_charge_m and column_charge_m must be >= 0")
    if not (0.0 <= d.drill_accuracy_sd_m < d.burden_m):
        raise ValueError("drill_accuracy_sd_m must satisfy 0 <= W < B")

    warnings: list[str] = []

    lb_lc = d.bottom_charge_m + d.column_charge_m
    if lb_lc > 0:
        diff_pct = abs(lb_lc - d.total_charge_m) / d.total_charge_m * 100.0
        if diff_pct > d.charge_sum_tolerance_pct:
            warnings.append(
                f"Lb + Lc = {lb_lc:.2f} m differs from Ltot = "
                f"{d.total_charge_m:.2f} m by {diff_pct:.1f} % "
                f"(> {d.charge_sum_tolerance_pct:.0f} % tolerance); check "
                "the charge length inputs"
            )

    if d.powder_factor_mode not in ("reported", "calculated"):
        raise ValueError("powder_factor_mode must be 'reported' or "
                         "'calculated'")
    if (d.powder_factor_mode == "reported"
            and d.powder_factor_reported_kg_m3 is None):
        raise ValueError("powder_factor_mode='reported' requires "
                         "powder_factor_reported_kg_m3")
    if (d.powder_factor_reported_kg_m3 is not None
            and d.powder_factor_reported_kg_m3 <= 0):
        raise ValueError("powder_factor_reported_kg_m3 must be > 0")

    return warnings


# ============================================================
# RESULT CONTAINER
# ============================================================
@dataclass
class KCOResult:
    """
    Full, traceable outcome of one KCO prediction.

    All final sizes are in millimetres. Every intermediate quantity that
    enters the prediction is recorded, together with the formulation and
    interpretation choices that produced it, so each number can be
    reproduced by hand.

    Attributes:
        design: The :class:`BlastDesign` used.
        rmd: Rock mass description rating used.
        jps: Joint plane spacing rating.
        jpa: Joint plane angle rating.
        jpa_tolerance_deg: Half-width of the 3-D JPA out/into cones used
            if the JPA was derived from the 3-D structural workflow.
        jcf: Joint condition factor, or None when the baseline
            JF = JPS + JPA form was used.
        jf: Joint factor.
        rdi: Rock density influence.
        hf: Hardness factor.
        bi: Blastability index.
        rock_factor: Rock factor A = 0.06 * BI.
        q_reported: Reported powder factor (kg/m3), or None.
        q_calculated: Powder factor from Q/(B*S*H) (kg/m3).
        q_used: Powder factor actually used in the X50 equation (kg/m3).
        q_difference_pct: |reported - calculated| / reported * 100, or
            None when no reported value exists.
        n: Cunningham (2005) uniformity index, Eq. (48).
        g_n: Shift factor actually applied (1.0 in the baseline).
        x50_mm: Median fragment size (mm).
        xmax_mm: Maximum fragment size (mm).
        xmax_governed_by: "in-situ block", "burden" or "spacing".
        block_size_method: Volume-to-length conversion used, or
            "user_defined".
        block_statistic: Statistic used on the DFN distribution, or None.
        block_percentile: Percentile used on the DFN distribution, or
            None.
        block_size_value_m: The in-situ block size that entered the xmax
            rule (m).
        b: Swebrec undulation parameter.
        warnings: All warnings raised during the prediction.
    """
    design: BlastDesign
    rmd: float
    jps: int
    jpa: int
    jpa_tolerance_deg: float
    jcf: Optional[float]
    jf: float
    rdi: float
    hf: float
    bi: float
    rock_factor: float
    q_reported: Optional[float]
    q_calculated: float
    q_used: float
    q_difference_pct: Optional[float]
    n: float
    g_n: float
    x50_mm: float
    xmax_mm: float
    xmax_governed_by: str
    block_size_method: str
    block_statistic: Optional[str]
    block_percentile: Optional[float]
    block_size_value_m: float
    b: float
    warnings: list = field(default_factory=list)

    # ---- distribution methods (KCO / Swebrec: the final model) ----
    def passing(self, x_mm) -> np.ndarray:
        """
        Return the KCO/Swebrec percentage passing at size(s) ``x_mm``.

        Args:
            x_mm: Fragment size(s) in millimetres.

        Returns:
            np.ndarray: Percentage passing in [0, 100].
        """
        return swebrec_passing(x_mm, self.x50_mm, self.xmax_mm, self.b)

    def size_at(self, passing_pct) -> np.ndarray:
        """
        Return the KCO/Swebrec fragment size(s) at given % passing.

        Args:
            passing_pct: Percentage passing in (0, 100].

        Returns:
            np.ndarray: Size(s) in millimetres.
        """
        return swebrec_size_at_passing(passing_pct, self.x50_mm,
                                       self.xmax_mm, self.b)

    def percentiles(self,
                    levels: Sequence[float] = (10, 20, 30, 50, 80, 90, 100)
                    ) -> dict[str, float]:
        """
        Return characteristic percentile sizes (mm).

        Defaults include X20, X50 and X80, which are the main quantities
        for comparing pre-blast and post-blast fragmentation.

        Args:
            levels: Percentage passing levels to evaluate.

        Returns:
            dict: e.g. ``{"X10": ..., "X20": ..., ..., "X100": ...}``.
        """
        return {f"X{int(p)}": float(self.size_at(p)) for p in levels}

    # ---- reference model, comparison only ----
    def passing_rosin_rammler(self, x_mm) -> np.ndarray:
        """
        Return the Rosin-Rammler (original Kuz-Ram) passing.

        REFERENCE MODEL ONLY — provided for comparison plots against the
        KCO/Swebrec prediction; not the final post-blast prediction.

        Args:
            x_mm: Fragment size(s) in millimetres.

        Returns:
            np.ndarray: Percentage passing in [0, 100].
        """
        return rosin_rammler_passing(x_mm, self.x50_mm, self.n)

    # ---- audit ----
    def audit_table(self) -> str:
        """
        Return a formatted audit table of every input and derived value.

        One row per quantity: parameter, symbol, value, unit and
        source/equation, so each site can be checked manually.

        Returns:
            str: Multi-line, fixed-width table.
        """
        d = self.design
        rows = [
            ("Burden", "B", f"{d.burden_m:.3g}", "m", "input"),
            ("Spacing", "S", f"{d.spacing_m:.3g}", "m", "input"),
            ("Bench height", "H", f"{d.bench_height_m:.3g}", "m", "input"),
            ("Hole diameter", "D", f"{d.hole_diameter_mm:.3g}", "mm",
             "input"),
            ("Drilling accuracy SD", "W", f"{d.drill_accuracy_sd_m:.3g}",
             "m", "input"),
            ("Bottom charge", "Lb", f"{d.bottom_charge_m:.3g}", "m",
             "input"),
            ("Column charge", "Lc", f"{d.column_charge_m:.3g}", "m",
             "input"),
            ("Total charge (above grade)", "Ltot",
             f"{d.total_charge_m:.3g}", "m", "input"),
            ("Charge per hole", "Q", f"{d.charge_per_hole_kg:.3g}", "kg",
             "input"),
            ("Powder factor (reported)", "q_rep",
             "-" if self.q_reported is None else f"{self.q_reported:.3g}",
             "kg/m3", "input"),
            ("Powder factor (calculated)", "q_calc",
             f"{self.q_calculated:.3g}", "kg/m3", "q = Q/(B*S*H)"),
            ("Powder factor (used)", "q", f"{self.q_used:.3g}", "kg/m3",
             f"mode = {d.powder_factor_mode}"),
            ("Weight strength vs ANFO", "s_ANFO", f"{d.s_anfo_pct:.3g}",
             "%", "input"),
            ("Rock density", "rho", f"{d.rock_density_kg_m3:.4g}",
             "kg/m3", "input"),
            ("UCS", "sigma_c", f"{d.ucs_mpa:.3g}", "MPa", "input"),
            ("Young's modulus", "E", f"{d.youngs_modulus_gpa:.3g}", "GPa",
             "input"),
            ("Mean joint spacing", "Sj", f"{d.mean_joint_spacing_m:.3g}",
             "m", "input (SPACING analysis)"),
            ("Joint plane spacing rating", "JPS", f"{self.jps}", "-",
             "jps_from_joint_spacing (see note 1)"),
            ("Joint plane angle rating", "JPA", f"{self.jpa}", "-",
             "input/derived"),
            ("JPA 3-D tolerance", "tol_JPA",
             f"{self.jpa_tolerance_deg:.1f}", "deg",
             "half-width of out/into cones; see note 2"),
            ("Joint condition factor", "JCF",
             "-" if self.jcf is None else f"{self.jcf:.2g}", "-",
             "not used (baseline)" if self.jcf is None else "input"),
            ("Joint factor", "JF", f"{self.jf:.4g}", "-",
             "JF = JPS + JPA" if self.jcf is None
             else "JF = JCF*JPS + JPA"),
            ("Rock mass case", "case", d.rock_mass_case, "-",
             "input"),
            ("Rock mass description", "RMD", f"{self.rmd:.4g}", "-",
             "from rock_mass_case"),
            ("Rock density influence", "RDI", f"{self.rdi:.4g}", "-",
             "RDI = 0.025*rho - 50"),
            ("Hardness factor", "HF", f"{self.hf:.4g}", "-",
             "HF = E/3 (E<50 GPa) else UCS/5"),
            ("Blastability index", "BI", f"{self.bi:.4g}", "-",
             "BI = RMD + RDI + HF"),
            ("Rock factor", "A", f"{self.rock_factor:.4g}", "-",
             "A = 0.06*BI"),
            ("Uniformity index", "n", f"{self.n:.4g}", "-",
             "Cunningham (2005) Eq. (48)"),
            ("Shift factor mode", "mode", d.shift_factor_mode, "-",
             "user choice"),
            ("Shift factor", "g(n)", f"{self.g_n:.4g}", "-",
             "g(n) = 1" if d.shift_factor_mode == "no_shift"
             else "mean-to-median shift"),
            ("Median size", "X50", f"{self.x50_mm:.4g}", "mm",
             "Kuznetsov / KCO eq. (11b)"),
            ("In-situ block size", "-", f"{self.block_size_value_m:.4g}",
             "m", f"method = {self.block_size_method}, "
             f"statistic = {self.block_statistic or self.block_percentile}"),
            ("Maximum size", "Xmax", f"{self.xmax_mm:.4g}", "mm",
             f"KCO eq. (11d), governed by {self.xmax_governed_by}"),
            ("Undulation parameter", "b", f"{self.b:.4g}", "-",
             "b = 2*ln2*n*ln(Xmax/X50)"),
        ]
        w = (28, 9, 12, 7, 40)
        header = (f"{'Parameter':<{w[0]}}{'Symbol':<{w[1]}}"
                  f"{'Value':>{w[2]}}  {'Unit':<{w[3]}}"
                  f"{'Source/Equation':<{w[4]}}")
        lines = [header, "-" * len(header)]
        for name, sym, val, unit, src in rows:
            lines.append(f"{name:<{w[0]}}{sym:<{w[1]}}{val:>{w[2]}}  "
                         f"{unit:<{w[3]}}{src:<{w[4]}}")
        lines.append("")
        lines.append("Notes (methodological assumptions to confirm):")
        lines.append(
            "  1. JPS classification: the literature leaves an ambiguity "
            "between\n     0.95*sqrt(B*S) and sqrt(B*S); using "
            "0.95*sqrt(B*S) as the lower\n     threshold of the JPS = 80 "
            "class is a selected boundary-handling\n     assumption."
        )
        lines.append(
            "  2. JPA 3-D tolerance = 45.0 degrees by default. This "
            "half-width of\n     the out-of-face and into-face cones "
            "around the face outward normal is a\n     selected modelling "
            "interpretation. It must be confirmed with the\n     supervisors "
            "before the JPA value is used in a final prediction."
        )
        if self.warnings:
            lines.append("")
            lines.append("Warnings:")
            for msg in self.warnings:
                lines.append(f"  - {msg}")
        return "\n".join(lines)


# ============================================================
# COMPARISON WITH MEASURED (e.g. WipFrag) FRAGMENTATION
# ============================================================
def compare_with_measured_fragmentation(result: KCOResult,
                                        measured_sizes_mm,
                                        measured_passing_pct) -> dict:
    """
    Compare a KCO prediction against measured post-blast fragmentation.

    Intended for WipFrag (or sieved) size distributions. Runs only when
    measured data are supplied; no data are ever fabricated. Measured
    percentile sizes (X20, X50, X80) are obtained by log-size /
    linear-passing (semi-log) interpolation of the measured curve:
    linear in cumulative passing percentage P, logarithmic in fragment
    size x. Where a percentile lies outside the measured passing range
    it is reported as NaN rather than extrapolated.

    Duplicate handling: if several measured points share the same
    passing percentage (a flat step in the curve), only the FIRST point
    (smallest size) of each run of equal P values is kept for the
    percentile interpolation, so np.interp operates on strictly
    increasing coordinates instead of behaving ambiguously. All points
    are still used for the RMSE/MAE/R^2 metrics.

    Input validation: sizes and passing values must be finite, sizes
    strictly positive and strictly increasing after sorting, passing
    values within [0, 100] and monotonically non-decreasing with
    increasing size (required because P is the interpolation
    coordinate).

    Metrics returned:
      - predicted and measured X20/X50/X80 (mm) with absolute and
        relative errors;
      - RMSE and MAE of percent passing, evaluated at the measured sizes;
      - R^2 of predicted vs measured percent passing.

    Args:
        result: A :class:`KCOResult` from :func:`predict_kco`.
        measured_sizes_mm: Measured fragment sizes (mm), monotonically
            increasing.
        measured_passing_pct: Measured cumulative percentage passing at
            those sizes, in [0, 100].

    Returns:
        dict: Metric name -> value.

    Raises:
        ValueError: If the arrays are empty, of different lengths, not
            finite, not strictly increasing in size, outside [0, 100]
            in passing, or not monotonically non-decreasing in passing.
    """
    x = np.asarray(measured_sizes_mm, dtype=float).ravel()
    p = np.asarray(measured_passing_pct, dtype=float).ravel()
    if x.size == 0 or x.size != p.size:
        raise ValueError("measured sizes and passing arrays must be "
                         "non-empty and of equal length")
    if not np.all(np.isfinite(x)) or not np.all(np.isfinite(p)):
        raise ValueError("measured sizes and passing values must all be "
                         "finite")
    if not np.all(x > 0):
        raise ValueError("measured sizes must all be > 0")
    if not np.all((p >= 0.0) & (p <= 100.0)):
        raise ValueError("measured passing values must lie in [0, 100]")
    order = np.argsort(x)
    x, p = x[order], p[order]
    if np.any(np.diff(x) <= 0):
        raise ValueError("measured sizes must be strictly increasing")
    if np.any(np.diff(p) < 0):
        raise ValueError(
            "measured_passing_pct must be monotonically non-decreasing "
            "with increasing fragment size"
        )

    pred_at_measured = result.passing(x)
    resid = pred_at_measured - p

    rmse = float(np.sqrt(np.mean(resid ** 2)))
    mae = float(np.mean(np.abs(resid)))
    ss_res = float(np.sum(resid ** 2))
    ss_tot = float(np.sum((p - p.mean()) ** 2))
    r2 = 1.0 - ss_res / ss_tot if ss_tot > 0 else float("nan")

    # Percentile interpolation coordinates: drop duplicate passing
    # values, keeping the first (smallest size) point of each run of
    # equal P, so the interpolation coordinate is strictly increasing.
    keep = np.concatenate(([True], np.diff(p) > 0))
    p_interp, x_interp = p[keep], x[keep]

    def _measured_size_at(level: float) -> float:
        # Log-size / linear-passing (semi-log) interpolation: linear in
        # P, logarithmic in x. Selected interpolation convention, not a
        # uniquely prescribed published algorithm.
        if level < p_interp.min() or level > p_interp.max():
            return float("nan")
        return float(np.exp(np.interp(level, p_interp, np.log(x_interp))))

    metrics: dict = {
        "rmse_passing_pct": rmse,
        "mae_passing_pct": mae,
        "r2_passing": r2,
        "n_measured_points": int(x.size),
    }
    for level in (20, 50, 80):
        pred = float(result.size_at(level))
        meas = _measured_size_at(float(level))
        metrics[f"x{level}_predicted_mm"] = pred
        metrics[f"x{level}_measured_mm"] = meas
        metrics[f"x{level}_error_mm"] = pred - meas
        metrics[f"x{level}_error_pct"] = (100.0 * (pred - meas) / meas
                                          if meas and not math.isnan(meas)
                                          else float("nan"))
    return metrics


# ============================================================
# END-TO-END PREDICTION
# ============================================================
def predict_kco(design: BlastDesign,
                block_volumes_m3=None) -> KCOResult:
    """
    Run the full KCO chain for one blast round.

    Steps (Cunningham 2005 formulation):
      1. joint terms and rock factor: JPS, JPA, (JCF), JF -> RMD -> BI
         -> A, with BI = RMD + RDI + HF and A = 0.06 * BI;
      2. powder factor cross-check (reported vs Q/(B*S*H));
      3. uniformity index n, Cunningham (2005) Eq. (48);
      4. median size X50, Kuznetsov, eq. (11b);
      5. in-situ block size from the DFN distribution using the
         explicit, user-selected method and statistic, then
         Xmax = min(block, B, S), eq. (11d);
      6. undulation parameter b = 2*ln2*n*ln(Xmax/X50), fully
         determining the Swebrec distribution of eq. (11a).

    Args:
        design: Fully populated :class:`BlastDesign`.
        block_volumes_m3: Optional iterable of DFN block volumes (m3).
            Required unless ``design.block_size_method`` is
            "user_defined", in which case
            ``design.in_situ_block_size_m`` is used directly.

    Returns:
        KCOResult: All intermediate quantities, choices, warnings and
            the callable size-distribution methods.

    Raises:
        ValueError: On invalid inputs, on inconsistent methodological
            selections, or when the resulting Xmax does not exceed X50.
    """
    warnings = _validate_design(design)

    # ---- 1. joint terms and rock factor ----
    jps = jps_from_joint_spacing(design.mean_joint_spacing_m,
                                 design.burden_m, design.spacing_m)

    if design.jpa_rating is not None:
        if design.jpa_rating not in (20, 30, 40):
            raise ValueError("jpa_rating must be 20, 30 or 40")
        jpa = design.jpa_rating
    elif design.jpa_case is not None:
        jpa = _joint_plane_angle_rating(design.jpa_case,
                                       design.jpa_mapping_version)
    else:
        raise ValueError("supply either jpa_case or jpa_rating")

    if design.jf_includes_jcf:
        if design.joint_condition is None:
            raise ValueError("jf_includes_jcf=True requires "
                             "joint_condition ('tight', 'relaxed' or "
                             "'gouge_filled')")
        jcf: Optional[float] = joint_condition_factor(design.joint_condition)
    else:
        jcf = None
    jf = joint_factor(jps, jpa, jcf)

    rmd = rock_mass_description(design.rock_mass_case, jf=jf)

    rdi = rock_density_influence(design.rock_density_kg_m3)
    hf = hardness_factor(design.youngs_modulus_gpa, design.ucs_mpa)
    bi = blastability_index(rmd, rdi, hf)
    a = rock_factor_A(bi)
    if a <= 0:
        raise ValueError(f"rock factor A = {a:.3g} is not positive; "
                         "check the BI terms")

    # ---- 2. powder factor: reported vs calculated ----
    q_calculated = calculate_powder_factor(design.charge_per_hole_kg,
                                           design.burden_m,
                                           design.spacing_m,
                                           design.bench_height_m)
    q_reported = design.powder_factor_reported_kg_m3
    q_difference_pct: Optional[float] = None
    if q_reported is not None:
        q_difference_pct = abs(q_reported - q_calculated) / q_reported * 100.0
        if q_difference_pct > design.q_tolerance_pct:
            warnings.append(
                f"reported powder factor ({q_reported:.3g} kg/m3) differs "
                f"from calculated Q/(B*S*H) ({q_calculated:.3g} kg/m3) "
                f"by {q_difference_pct:.1f} % "
                f"(> {design.q_tolerance_pct:.0f} % tolerance); the "
                f"'{design.powder_factor_mode}' value is used, the "
                "reported value is never overwritten automatically"
            )
    q_used = (q_reported if design.powder_factor_mode == "reported"
              else q_calculated)

    # ---- 3. uniformity index ----
    if design.timing_scatter_factor_ns is None:
        raise ValueError(
            "timing_scatter_factor_ns is required for Cunningham 2005 n; "
            "set it explicitly or compute it from in-row delay scatter"
        )
    n = uniformity_index_n(
        burden_m=design.burden_m,
        spacing_m=design.spacing_m,
        hole_diameter_mm=design.hole_diameter_mm,
        drill_accuracy_sd_m=design.drill_accuracy_sd_m,
        charge_length_m=design.total_charge_m,
        bench_height_m=design.bench_height_m,
        rock_factor_A=a,
        timing_scatter_factor_ns=design.timing_scatter_factor_ns,
    )
    if n < 0.6 or n > 2.2:
        warnings.append(
            f"uniformity index n = {n:.2f} lies outside the typical "
            "0.6-2.2 range (Cunningham); check geometry inputs and units "
            "(B in m, D in mm)"
        )

    # ---- 4. median size ----
    g_n = shift_factor_g(n, mode=design.shift_factor_mode)
    x50_cm = x50_kuznetsov(
        rock_factor_a=a,
        charge_per_hole_kg=design.charge_per_hole_kg,
        powder_factor_kg_m3=q_used,
        s_anfo_pct=design.s_anfo_pct,
        g_n=g_n,
        timing_factor=design.timing_factor,
    )
    x50_mm = x50_cm * 10.0  # cm -> mm, the only unit conversion applied

    # ---- 5. in-situ block size and Xmax ----
    if design.block_size_method == "user_defined":
        if (design.in_situ_block_size_m is None
                or design.in_situ_block_size_m <= 0):
            raise ValueError("block_size_method='user_defined' requires a "
                             "positive in_situ_block_size_m")
        block_size_m = design.in_situ_block_size_m
        block_statistic = design.block_statistic
        block_percentile = design.block_percentile
    else:
        if block_volumes_m3 is None:
            raise ValueError(
                f"block_size_method='{design.block_size_method}' requires "
                "the DFN block_volumes_m3 argument"
            )
        block_size_m, info = characteristic_block_size_from_distribution(
            block_volumes_m3,
            block_size_method=design.block_size_method,
            block_statistic=design.block_statistic,
            block_percentile=design.block_percentile,
        )
        warnings.extend(info["warnings"])
        block_statistic = info["block_statistic"]
        block_percentile = info["block_percentile"]

    xmax_m, governed_by = xmax_kco(block_size_m, design.burden_m,
                                   design.spacing_m)
    xmax_mm = xmax_m * 1000.0

    if xmax_mm <= x50_mm:
        raise ValueError(
            f"inconsistent inputs: predicted X50 = {x50_mm:.0f} mm is not "
            f"smaller than Xmax = {xmax_mm:.0f} mm (governed by "
            f"{governed_by}). Check the powder factor, the rock factor A "
            "and the in-situ block size."
        )
    if xmax_mm < 1.5 * x50_mm:
        warnings.append(
            f"Xmax ({xmax_mm:.0f} mm) is only slightly larger than X50 "
            f"({x50_mm:.0f} mm). Because b = 2*ln(2)*n*ln(Xmax/X50), this "
            "may produce an unusually small b and a strongly constrained "
            "Swebrec distribution. Verify Xmax, X50 and the underlying "
            "inputs."
        )

    # ---- 6. undulation parameter ----
    b = b_parameter(xmax_mm, x50_mm, n)
    if b < B_TYPICAL_MIN or b > B_TYPICAL_MAX:
        warnings.append(
            f"Calculated b = {b:.2f} is outside the typical Swebrec range "
            f"{B_TYPICAL_MIN:.0f}-{B_TYPICAL_MAX:.0f}. This does not "
            "automatically mean the result is invalid, but the inputs and "
            "units should be checked."
        )

    return KCOResult(
        design=design,
        rmd=rmd,
        jps=jps,
        jpa=jpa,
        jpa_tolerance_deg=design.jpa_tolerance_deg,
        jcf=jcf,
        jf=jf,
        rdi=rdi,
        hf=hf,
        bi=bi,
        rock_factor=a,
        q_reported=q_reported,
        q_calculated=q_calculated,
        q_used=q_used,
        q_difference_pct=q_difference_pct,
        n=n,
        g_n=g_n,
        x50_mm=x50_mm,
        xmax_mm=xmax_mm,
        xmax_governed_by=governed_by,
        block_size_method=design.block_size_method,
        block_statistic=block_statistic,
        block_percentile=block_percentile,
        block_size_value_m=block_size_m,
        b=b,
        warnings=warnings,
    )


# ============================================================
# STRUCTURAL / DFN INPUTS FOR JPA
# ============================================================
@dataclass
class StructuralFamily:
    """
    One fracture family from the existing DFN / structural workflow.

    Reuses the data already produced by the VARENNE pipeline; no manual
    dip/dipdir re-entry is required.
    """
    name: str
    dip_deg: float
    dipdir_deg: float
    p32: Optional[float] = None
    count: Optional[int] = None
    total_area_m2: Optional[float] = None


@dataclass
class StructuralInputs:
    """
    Existing structural/DFN data required for the JPA workflow.

    Built by :func:`load_varenne_structural_inputs` from the project's
    existing files:

      - ``run_site.py`` ``SITE_CONFIGS[site]``: ``dip_face`` and
        ``dipdir_face`` of the free/mapping surface;
      - ``DFN_fracture_characteristics_VARENNE.csv``: every generated
        fracture's ``family_name``, ``dip_deg``, ``dipdir_deg`` and
        ``area_m2``;
      - ``P32_calibrated_summary.csv``: the existing family-intensity
        measure ``P32_calibrated`` used for dominant-family selection.
    """
    site: str
    face_dip_deg: float
    face_dipdir_deg: float
    families: list = field(default_factory=list)

def _parse_run_site_face(site_name: str = "VARENNE",
                         run_site_path: Optional[str] = None
                         ) -> tuple[float, float]:
    """
    Read ``dip_face`` and ``dipdir_face`` for a site from ``run_site.py``.

    Regex parsing is used deliberately: importing ``run_site.py`` would
    execute its side effects (UnBlocks-Gen banner and possibly the whole
    pipeline). The values are the same as the existing ``SITE_CONFIGS``
    dictionary.

    Args:
        site_name: Site key as used in ``SITE_CONFIGS``.
        run_site_path: Path to ``run_site.py``. Defaults to the file in
            the same directory as this module.

    Returns:
        tuple[float, float]: (dip_face, dipdir_face) in degrees.

    Raises:
        ValueError: If the site or the face values are not found.
    """
    if run_site_path is None:
        run_site_path = os.path.join(os.path.dirname(__file__), "run_site.py")
    with open(run_site_path, "r", encoding="utf-8") as f:
        text = f.read()

    pat = rf'"{re.escape(site_name)}"\s*:\s*\{{'
    m = re.search(pat, text)
    if not m:
        raise ValueError(f"site {site_name!r} not found in {run_site_path}")

    # Extract the opening brace and the matching closing brace.
    start = m.end() - 1
    depth = 0
    end = start
    for i in range(start, len(text)):
        if text[i] == '{':
            depth += 1
        elif text[i] == '}':
            depth -= 1
            if depth == 0:
                end = i + 1
                break
    block = text[start:end]

    dip_m = re.search(r'"dip_face"\s*:\s*([0-9.]+)', block)
    dipdir_m = re.search(r'"dipdir_face"\s*:\s*([0-9.]+)', block)
    if not dip_m or not dipdir_m:
        raise ValueError(
            f"dip_face / dipdir_face not found in site block for "
            f"{site_name!r} in {run_site_path}"
        )
    return float(dip_m.group(1)), float(dipdir_m.group(1))


def _dip_dipdir_to_upper_normal(dip_deg: float, dipdir_deg: float) -> tuple:
    """
    Convert dip/dip-direction to the upward-pointing unit normal vector.

    The normal of a plane with dip ``d`` and dip direction ``dd`` has
    trend = dd and plunge = 90° - d. Resolving this pole into a unit
    vector gives:

        x = sin(dd) * sin(d)
        y = cos(dd) * sin(d)
        z = cos(d)

    This vector always points to the upper hemisphere (z >= 0 for
    0 <= dip <= 90°).
    """
    d = math.radians(dip_deg)
    dd = math.radians(dipdir_deg)
    sin_d = math.sin(d)
    return (
        math.sin(dd) * sin_d,
        math.cos(dd) * sin_d,
        math.cos(d),
    )


def _normal_to_dip_dipdir(nx: float, ny: float, nz: float) -> tuple[float, float]:
    """
    Convert a unit normal (upper hemisphere) back to dip and dip direction.

    dip    = arccos(nz)
    dipdir = atan2(nx, ny)  (because x/y = tan(dd) for a pole)
    """
    # Force upper hemisphere for the returned plane.
    if nz < 0:
        nx, ny, nz = -nx, -ny, -nz
    nz = max(-1.0, min(1.0, nz))
    dip = math.degrees(math.acos(nz))
    dipdir = math.degrees(math.atan2(nx, ny)) % 360.0
    return dip, dipdir


def _mean_plane_from_measurements(measurements: list[tuple[float, float]]
                                  ) -> tuple[float, float, int]:
    """
    Compute a representative dip/dipdir by averaging 3-D plane-normal vectors.

    Each measurement is converted to its upper-hemisphere unit normal. The
    mean vector (vector mean, i.e. Fisher's spherical mean) is then
    normalised and converted back to a dip and dip direction. This avoids
    the circular-statistics problem that a naive arithmetic mean of dip
    directions would create near the 0/360° discontinuity.

    Args:
        measurements: List of (dip_deg, dipdir_deg) for one set.

    Returns:
        tuple[float, float, int]: (mean_dip, mean_dipdir, n).
    """
    if not measurements:
        raise ValueError("no measurements supplied for mean-plane calculation")
    sx = sy = sz = 0.0
    for dip_deg, dipdir_deg in measurements:
        nx, ny, nz = _dip_dipdir_to_upper_normal(dip_deg, dipdir_deg)
        sx += nx
        sy += ny
        sz += nz
    norm = math.sqrt(sx * sx + sy * sy + sz * sz)
    if norm == 0.0:
        raise ValueError("measurements are too dispersed to form a mean plane")
    mean_dip, mean_dipdir = _normal_to_dip_dipdir(sx / norm, sy / norm, sz / norm)
    return mean_dip, mean_dipdir, len(measurements)


def _load_measured_family_orientations(measured_excel_path: Optional[str] = None
                                       ) -> dict:
    """
    Load the original measured discontinuities from DIPSVARENNE.xlsx.

    The file has columns ``Dip``, ``Dip Direction`` and ``Corrected Set``.
    ``Corrected Set`` values 1..4 are mapped to family names ``fam1``..``fam4``.
    For each set a representative dip/dipdir is computed from the 3-D normal
    vector mean of all the set's measurements. The count of valid
    measurements is also returned.

    Args:
        measured_excel_path: Path to DIPSVARENNE.xlsx. Defaults to
            ``assets/DIPSVARENNE.xlsx``.

    Returns:
        dict: ``{family_name: {dip_deg, dipdir_deg, count}}``.
    """
    if measured_excel_path is None:
        measured_excel_path = os.path.join(
            os.path.dirname(__file__), "assets", "DIPSVARENNE.xlsx")
    wb = openpyxl.load_workbook(measured_excel_path, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError(f"no active sheet in {measured_excel_path}")

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    try:
        dip_col = headers.index("Dip") + 1
        dipdir_col = headers.index("Dip Direction") + 1
        set_col = headers.index("Corrected Set") + 1
    except ValueError as exc:
        raise ValueError(
            f"required column missing in {measured_excel_path}; "
            f"expected 'Dip', 'Dip Direction' and 'Corrected Set' "
            f"but got {headers}"
        ) from exc

    raw: dict = {}
    for r in range(2, ws.max_row + 1):
        dip = ws.cell(r, dip_col).value
        dipdir = ws.cell(r, dipdir_col).value
        s = ws.cell(r, set_col).value
        if dip is None or dipdir is None or s is None:
            continue
        try:
            s_int = int(float(s))
        except (TypeError, ValueError):
            continue
        if s_int not in (1, 2, 3, 4):
            continue
        fam = f"fam{s_int}"
        if fam not in raw:
            raw[fam] = []
        raw[fam].append((float(dip), float(dipdir)))

    out = {}
    for fam in sorted(raw):
        dip, dipdir, n = _mean_plane_from_measurements(raw[fam])
        out[fam] = {"dip_deg": dip, "dipdir_deg": dipdir, "count": n,
                    "total_area_m2": None}
    return out


def _load_dfn_family_orientations(dfn_csv: Optional[str] = None) -> dict:
    """
    Aggregate per-family dip/dipdir from the DFN fracture characteristics CSV.

    Returns a dict ``{family_name: {dip_deg, dipdir_deg, count,
    total_area_m2}}``. The family means are arithmetic means of all
    generated DFN fractures for that family.

    Args:
        dfn_csv: Path to the DFN characteristics CSV. Defaults to
            ``outputs/combined/DFN_fracture_characteristics_VARENNE.csv``.

    Returns:
        dict: Aggregated orientations and counts/areas per family.
    """
    if dfn_csv is None:
        dfn_csv = os.path.join(
            os.path.dirname(__file__), "outputs", "combined",
            "DFN_fracture_characteristics_VARENNE.csv")
    by_fam: dict = {}
    with open(dfn_csv, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            fam = row.get("family_name")
            if not fam:
                continue
            if fam not in by_fam:
                by_fam[fam] = {"dips": [], "dipdirs": [], "areas": []}
            by_fam[fam]["dips"].append(float(row["dip_deg"]))
            by_fam[fam]["dipdirs"].append(float(row["dipdir_deg"]))
            by_fam[fam]["areas"].append(float(row["area_m2"]))

    out = {}
    for fam, d in by_fam.items():
        out[fam] = {
            "dip_deg": float(np.mean(d["dips"])),
            "dipdir_deg": float(np.mean(d["dipdirs"])),
            "count": len(d["dips"]),
            "total_area_m2": float(np.sum(d["areas"])),
        }
    return out


def _load_p32_weights(p32_csv: Optional[str] = None) -> dict:
    """
    Load the existing calibrated P32 per family.

    Args:
        p32_csv: Path to ``P32_calibrated_summary.csv``. Defaults to
            ``outputs/VARENNE/02_calibration/P32_calibrated_summary.csv``.

    Returns:
        dict: ``{family_name: P32_calibrated}``.
    """
    if p32_csv is None:
        p32_csv = os.path.join(
            os.path.dirname(__file__), "outputs", "VARENNE",
            "02_calibration", "P32_calibrated_summary.csv")
    weights = {}
    with open(p32_csv, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            fam = f"fam{row['fam']}"
            weights[fam] = float(row["P32_calibrated"])
    return weights


def load_varenne_structural_inputs(
        site_name: str = "VARENNE",
        measured_excel_path: Optional[str] = None,
        dfn_csv: Optional[str] = None,
        p32_csv: Optional[str] = None,
        run_site_path: Optional[str] = None,
        use_measured_excel: bool = True) -> StructuralInputs:
    """
    Build a :class:`StructuralInputs` object from existing project files.

    This is the main integration point: it reuses the face orientation in
    ``run_site.py`` and the per-family orientations and calibrated P32
    values that already exist in the project, so the user does not re-enter
    dip/dipdir values.

    By default the family orientations come from the original measured
    discontinuities in ``DIPSVARENNE.xlsx`` (``Corrected Set`` 1..4), with
    a 3-D normal-vector mean plane computed per set. The DFN-generated
    fracture CSV is available as an optional fallback.

    Args:
        site_name: Site key as used in ``run_site.py`` (default "VARENNE").
        measured_excel_path: Optional override for the DIPSVARENNE.xlsx
            file. Defaults to ``assets/DIPSVARENNE.xlsx``.
        dfn_csv: Optional override for the DFN characteristics CSV; only
            used when ``use_measured_excel=False``.
        p32_csv: Optional override for the P32 calibration CSV.
        run_site_path: Optional override for ``run_site.py``.
        use_measured_excel: If True (default), use the original measured
            Excel. If False, use the generated DFN CSV.

    Returns:
        StructuralInputs: Face orientation plus the family list.
    """
    face_dip, face_dipdir = _parse_run_site_face(site_name, run_site_path)
    if use_measured_excel:
        orients = _load_measured_family_orientations(measured_excel_path)
    else:
        orients = _load_dfn_family_orientations(dfn_csv)
    p32 = _load_p32_weights(p32_csv)
    families = []
    for fam in sorted(orients):
        o = orients[fam]
        families.append(StructuralFamily(
            name=fam,
            dip_deg=o["dip_deg"],
            dipdir_deg=o["dipdir_deg"],
            p32=p32.get(fam),
            count=o["count"],
            total_area_m2=o["total_area_m2"],
        ))
    return StructuralInputs(
        site=site_name,
        face_dip_deg=face_dip,
        face_dipdir_deg=face_dipdir,
        families=families,
    )


def calculate_jpa_for_site(site_name: str = "VARENNE",
                           jpa_mapping_version: str = "cunningham_2005",
                           dominant_by: Optional[str] = "p32_calibrated",
                           subhorizontal_dip_deg: float = 30.0,
                           tolerance_deg: float = 45.0) -> dict:
    """
    Compute the per-family JPA classification for one site.

    Loads the rock-face orientation and the measured family orientations,
    then compares the full 3-D joint and face planes for every family.
    Returns the per-family table, the dominant family chosen by the
    selected criterion, and the final JPA used by KCO.

    This is the single active JPA entry point. The old 2-D helpers
    (``jpa_from_orientation``, ``face_azimuth_deg``) have been removed;
    all JPA now flows through this 3-D function.

    Args:
        site_name: Site key in ``run_site.py`` ``SITE_CONFIGS``.
        jpa_mapping_version: "cunningham_2005" (baseline) or
            "cunningham_1987".
        dominant_by: "p32_calibrated", "count", "total_area_m2", or
            ``None``.
        subhorizontal_dip_deg: Dip below which a set is treated as sub-
            horizontal (default 30°).
        tolerance_deg: Angular half-width of the "out of face" and
            "into face" cones around the face outward normal (default 45°).

    Returns:
        dict with keys ``families``, ``dominant_family``,
        ``selection_method``, ``final_jpa`` and ``warnings``.
    """
    structural_inputs = load_varenne_structural_inputs(site_name=site_name)
    families = []
    warnings: list[str] = []

    for fam in structural_inputs.families:
        jpa, case, geometry = _jpa_from_3d_planes(
            joint_dip_deg=fam.dip_deg,
            joint_dipdir_deg=fam.dipdir_deg,
            face_dip_deg=structural_inputs.face_dip_deg,
            face_dipdir_deg=structural_inputs.face_dipdir_deg,
            subhorizontal_dip_deg=subhorizontal_dip_deg,
            tolerance_deg=tolerance_deg,
            jpa_mapping_version=jpa_mapping_version,
        )
        families.append({
            "family": fam.name,
            "dip_deg": fam.dip_deg,
            "dipdir_deg": fam.dipdir_deg,
            "face_dip_deg": structural_inputs.face_dip_deg,
            "face_dipdir_deg": structural_inputs.face_dipdir_deg,
            "classification": case,
            "jpa": jpa,
            "alpha_deg": geometry.get("alpha_deg"),
            "p32_calibrated": fam.p32,
            "count": fam.count,
            "total_area_m2": fam.total_area_m2,
            "geometry": geometry,
        })

    dominant_family = None
    if dominant_by is None:
        warnings.append(
            "No dominant-family selection criterion was supplied. "
            "The JPA of every family is reported; the final family for "
            "the KCO prediction still needs a methodological decision or "
            "a manual override (BlastDesign.jpa_case / jpa_rating)."
        )
    else:
        if dominant_by not in ("p32_calibrated", "count", "total_area_m2"):
            raise ValueError(
                f"unknown dominant_by {dominant_by!r}; expected "
                "'p32_calibrated', 'count', 'total_area_m2' or None"
            )

        missing = [r for r in families if r[dominant_by] is None]
        if missing:
            warnings.append(
                f"Cannot select a dominant family by {dominant_by}: "
                f"missing the selected measure for "
                f"{', '.join(m['family'] for m in missing)}. "
                f"Choose a measure that exists for all families or "
                f"set dominant_by=None."
            )
        else:
            dom = max(families, key=lambda r: r[dominant_by])
            dominant_family = {
                "family": dom["family"],
                "classification": dom["classification"],
                "jpa": dom["jpa"],
                "weight_value": dom[dominant_by],
            }

    final_jpa = None
    if dominant_family is not None:
        final_jpa = dominant_family["jpa"]

    return {
        "families": families,
        "dominant_family": dominant_family,
        "selection_method": dominant_by,
        "final_jpa": final_jpa,
        "warnings": warnings,
    }


def format_jpa_result(result: dict,
                       include_weights: bool = True) -> str:
    """
    Return a printable table of the per-family JPA classification.

    Args:
        result: Output of :func:`calculate_jpa_for_site`.
        include_weights: Also print P32, count and total area.

    Returns:
        str: Fixed-width table.
    """
    header = (
        f"{'Family':<7} | {'Dip':>6} | {'Dipdir':>7} | "
        f"{'Face dip':>9} | {'Face dipdir':>12} | "
        f"{'Alpha':>7} | {'Classification':<28} | {'JPA':>4}"
    )
    if include_weights:
        header += (
            f" | {'P32':>7} | {'Count':>6} | {'Area m2':>9}"
        )
    lines = [header, "-" * len(header)]
    for r in result["families"]:
        alpha = r["alpha_deg"]
        alpha_s = f"{alpha:>7.1f}" if alpha is not None else "      -"
        s = (
            f"{r['family']:<7} | {r['dip_deg']:>6.1f} | "
            f"{r['dipdir_deg']:>7.1f} | {r['face_dip_deg']:>9.1f} | "
            f"{r['face_dipdir_deg']:>12.1f} | "
            f"{alpha_s} | {r['classification']:<28} | {r['jpa']:>4}"
        )
        if include_weights:
            p32 = r["p32_calibrated"]
            p32_s = f"{p32:>7.4f}" if p32 is not None else "    -  "
            area = r["total_area_m2"]
            area_s = f"{area:>9.1f}" if area is not None else "      -  "
            s += (f" | {p32_s} | {r['count']:>6} | {area_s}")
        lines.append(s)
    if result["dominant_family"] is not None:
        d = result["dominant_family"]
        selection = result.get("selection_method") or "unknown"
        lines.append("")
        lines.append(
            f"Dominant family (by {selection}): "
            f"{d['family']}  final JPA = {d['jpa']} "
            f"({d['classification']})"
        )
    if result["final_jpa"] is not None:
        lines.append("")
        lines.append(f"Final JPA for KCO: {result['final_jpa']}")
    if result["warnings"]:
        lines.append("")
        lines.append("Warnings:")
        for w in result["warnings"]:
            lines.append(f"  - {w}")
    return "\n".join(lines)


# ============================================================
# SELF-TEST — Bårarp round 4, worked example of Ouchterlony (2005)
# ============================================================
def self_test(verbose: bool = True) -> bool:
    """
    Sanity check of the active Cunningham 2005 KCO workflow.

    Verifies that the key equations evaluate without error and are
    internally consistent. The X50 reference is the Bårarp Round 4 worked
    example of Ouchterlony (2005); the uniformity index n is now computed
    with the Cunningham (2005) / Ouchterlony & Sanchidrian (2019)
    Eq. (48) form, so its numerical value is no longer compared to the
    legacy 1987 example.

    Inputs: A = 13, Q = 9.24 kg, q = 0.55 kg/m3,
    s_ANFO = 62.2 %, D = 51 mm, B = 1.8 m, S = 2.2 m, H = 5.2 m,
    Ltot = 3.9 m above grade, Lb = Ltot, Lc = 0, W = 0.25 m,
    xmax = 2000 mm.

    Checks performed: X50, P(X50) = 50 %, P(Xmax) = 100 %,
    inverse-Swebrec round trip, and the cm -> mm unit conversion.

    Args:
        verbose: Print the comparison table.

    Returns:
        bool: True if every check passes within tolerance.
    """
    n = uniformity_index_n(
        burden_m=1.8, spacing_m=2.2,
        hole_diameter_mm=51.0, drill_accuracy_sd_m=0.25,
        charge_length_m=3.9, bench_height_m=5.2,
        rock_factor_A=13.0, timing_scatter_factor_ns=1.0)
    x50_cm = x50_kuznetsov(rock_factor_a=13.0, charge_per_hole_kg=9.24,
                           powder_factor_kg_m3=0.55, s_anfo_pct=62.2)
    b = b_parameter(xmax=2000.0, x50=x50_cm * 10.0, n=n)

    checks = [
        ("X50 (cm)", x50_cm, 44.8, 0.1),
    ]

    ok = True
    if verbose:
        print("KCO self-test -- active Cunningham 2005 workflow sanity "
              "check\n(Ouchterlony & Sanchidrian 2019, Eq. 48). "
              "This checks internal consistency,\nnot the model's validity "
              "for any particular site.\n")
        print(f"{'quantity':<14}{'computed':>12}{'reference':>12}{'':>8}")
    for label, got, expected, tol in checks:
        passed = abs(got - expected) <= tol
        ok = ok and passed
        if verbose:
            print(f"{label:<14}{got:>12.4f}{expected:>12.4f}"
                  f"{'  OK' if passed else '  FAIL':>8}")

    x50_mm = x50_cm * 10.0

    # Swebrec fixed points.
    p50 = float(swebrec_passing(x50_mm, x50_mm, 2000.0, b))
    p_max = float(swebrec_passing(2000.0, x50_mm, 2000.0, b))
    for label, got, expected in (("P(X50) %", p50, 50.0),
                                 ("P(Xmax) %", p_max, 100.0)):
        passed = abs(got - expected) < 1e-6
        ok = ok and passed
        if verbose:
            print(f"{label:<14}{got:>12.4f}{expected:>12.4f}"
                  f"{'  OK' if passed else '  FAIL':>8}")

    # Inverse round trip.
    x_at_30 = float(swebrec_size_at_passing(30.0, x50_mm, 2000.0, b))
    back = float(swebrec_passing(x_at_30, x50_mm, 2000.0, b))
    passed = abs(back - 30.0) < 1e-6
    ok = ok and passed
    if verbose:
        print(f"{'invert P=30':<14}{back:>12.4f}{30.0:>12.4f}"
              f"{'  OK' if passed else '  FAIL':>8}")

    # Unit conversion: the Kuznetsov equation returns cm; final outputs
    # are mm.
    passed = abs(x50_mm - 448.373) < 0.5
    ok = ok and passed
    if verbose:
        print(f"{'cm->mm conv.':<14}{x50_mm:>12.1f}{448.4:>12.1f}"
              f"{'  OK' if passed else '  FAIL':>8}")
        print("\nAll checks passed: the implementation reproduces the "
              "published example\nwithin numerical tolerance."
              if ok else "\nSOME CHECKS FAILED.")
    return ok


if __name__ == "__main__":
    raise SystemExit(0 if self_test() else 1)
