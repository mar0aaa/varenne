import os
import sys
import numpy as np

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPT_DIR)

from run_site import SITE_CONFIGS, run_site
from extract_dfn_diameters import export_dfn_fracture_characteristics

MENU_ITEMS = {
    # ──── DFN + Blockometry ────
    "1": ("VARENNE",      "VARENNE     — DFN calibration + blockometry (fam1, fam2, fam3, fam4)"),

    # ──── Persistence / Spacing ────
    "2": ("PERSISTENCE",  "PERSISTENCE — Trace length survival fits"),
    "3": ("SPACING",      "SPACING     — Perpendicular spacing survival fits"),

    # ──── Exports ────
    "4": ("DFNCHAR",      "Export DFN fracture characteristics (dip, dipdir, area, diameter)"),
    "5": ("BLOCKOMETRY",  "Blockometry percentile summary"),
    "6": ("TRACELENGTHS", "Export trace length Excel: DFN vs CloudCompare"),
    "7": ("PLOTVOLUMES",  "Plot block volume distribution"),
}


def _run_persistence():
    import importlib.util
    spec = importlib.util.spec_from_file_location("PERSISTENCE", os.path.join(SCRIPT_DIR, "PERSISTENCE.py"))
    mod  = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    mod.main()


def _run_spacing():
    import importlib.util
    spec = importlib.util.spec_from_file_location("SPACING", os.path.join(SCRIPT_DIR, "SPACING.py"))
    mod  = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)


def _run_blockometry():
    import importlib.util
    spec = importlib.util.spec_from_file_location(
        "export_blockometry_charts",
        os.path.join(SCRIPT_DIR, "export_blockometry_charts.py")
    )
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    mod.main()


def _run_plotvolumes():
    import importlib.util
    spec = importlib.util.spec_from_file_location(
        "PLOTVARENNE",
        os.path.join(SCRIPT_DIR, "PLOTVARENNE.py")
    )
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)


def _export_trace_length_excel():
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    assets = os.path.join(SCRIPT_DIR, "assets")
    trace_file = SITE_CONFIGS["VARENNE"]["trace_name"]
    region_filter = SITE_CONFIGS["VARENNE"]["region_filter"]
    persist_dir = os.path.join(SCRIPT_DIR, "outputs", "VARENNE", "07_persistence")

    thin   = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    row_fill = PatternFill("solid", fgColor="DEEAF1")

    def _style_header(ws, n):
        for c in range(1, n + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = center; cell.border = border
            ws.column_dimensions[get_column_letter(c)].width = 18

    def _style_row(ws, r, n):
        for c in range(1, n + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = row_fill; cell.alignment = center; cell.border = border

    wb = openpyxl.Workbook()

    # Sheet 1 — CC trace lengths
    ws1 = wb.active; ws1.title = "CC Trace Lengths"
    ws1.append(["Site", "Family", "Length (m)"])
    _style_header(ws1, 3)
    path = os.path.join(assets, trace_file)
    if os.path.exists(path):
        df = pd.read_csv(path)
        if region_filter and "REGION" in df.columns:
            df = df[df["REGION"] == region_filter]
        df["Length"] = pd.to_numeric(df["Length"], errors="coerce")
        df = df.dropna(subset=["fam", "Length"])
        df = df[df["Length"] > 0]
        for _, row in df.iterrows():
            ws1.append(["VARENNE", str(row["fam"]), round(float(row["Length"]), 4)])
            _style_row(ws1, ws1.max_row, 3)

    # Sheet 2 — DFN trace lengths
    ws2 = wb.create_sheet("DFN Trace Lengths")
    ws2.append(["Site", "Family", "Radius (m)", "Diameter (m)"])
    _style_header(ws2, 4)
    if os.path.isdir(persist_dir):
        for fname in sorted(os.listdir(persist_dir)):
            if fname.startswith("persistence_radii_") and fname.endswith(".csv"):
                fam = fname.replace("persistence_radii_", "").replace(".csv", "")
                ddf = pd.read_csv(os.path.join(persist_dir, fname))
                for _, row in ddf.iterrows():
                    r = round(float(row.get("radius_m", float("nan"))), 4)
                    d = round(float(row.get("diameter_m", float("nan"))), 4)
                    ws2.append(["VARENNE", fam, r, d])
                    _style_row(ws2, ws2.max_row, 4)

    out = os.path.join(SCRIPT_DIR, "outputs", "trace_length_comparison.xlsx")
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    print(f"\n✅ Saved: {out}")


def main():
    print("\n" + "=" * 60)
    print("  VARENNE — DFN Analysis Menu")
    print("=" * 60)
    for key in sorted(MENU_ITEMS, key=lambda x: int(x)):
        print(f"  [{key}] {MENU_ITEMS[key][1]}")
    print("  [0] Exit")
    print("=" * 60)

    choice = input("Enter choice: ").strip()

    if choice == "0":
        print("Goodbye.")
        return

    if choice not in MENU_ITEMS:
        print(f"Unknown option: {choice}")
        return

    action = MENU_ITEMS[choice][0]

    if action == "VARENNE":
        run_site(SITE_CONFIGS["VARENNE"])

    elif action == "PERSISTENCE":
        _run_persistence()

    elif action == "SPACING":
        _run_spacing()

    elif action == "DFNCHAR":
        out_dir = os.path.join(SCRIPT_DIR, "outputs", "VARENNE", "07_persistence")
        export_dfn_fracture_characteristics(
            dfn_vtk_dir=os.path.join(SCRIPT_DIR, "outputs", "VARENNE", "03_dfn_vtk"),
            output_dir=out_dir,
            site_name="VARENNE",
        )

    elif action == "BLOCKOMETRY":
        _run_blockometry()

    elif action == "TRACELENGTHS":
        _export_trace_length_excel()

    elif action == "PLOTVOLUMES":
        _run_plotvolumes()

    else:
        print(f"Action '{action}' not yet implemented.")


if __name__ == "__main__":
    main()
